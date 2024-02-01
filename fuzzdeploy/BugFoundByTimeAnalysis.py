import csv
import json
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

from . import utility
from .CasrTriageAnalysis import CasrTriageAnalysis
from .constants import *
from .ExcelManager import ExcelManager


class BugFoundByTimeAnalysis:
    fuzzer_colors = {
        "aflplusplus": "00A0B0",
        "htfuzz": "FBD26A",
    }
    display_fields = [
        FUZZER,
        REPEAT,
    ]

    @staticmethod
    def get_csv_data(file_path):
        crash_to_time = {}
        last_index = 0
        base_time = None
        with open(
            file_path,
            newline="",
        ) as csvfile:
            csv_reader = csv.DictReader(csvfile)
            for row in csv_reader:
                relative_time = None
                crash_index = None
                for key, val in row.items():
                    if "unix_time" in key or "relative_time" in key:
                        relative_time = int(val)
                        if "unix_time" in key:
                            if not base_time:
                                base_time = relative_time
                                relative_time = 0
                            else:
                                relative_time -= base_time
                    if "unique_crashes" in key or "saved_crashes" in key:
                        crash_index = int(val)
                assert (
                    relative_time is not None and crash_index is not None
                ), "invalid row"
                if crash_index > last_index:
                    last_index = crash_index
                    crash_to_time[crash_index] = relative_time
        return crash_to_time

    @staticmethod
    def obtain(WORK_DIR):
        assert os.path.exists(WORK_DIR), f"{WORK_DIR} not exists"
        assert os.path.exists(
            os.path.join(WORK_DIR, TRIAGE_BY_CASR)
        ), "Please make crash triage first"
        bug_info = {}
        for ar_path in utility.get_workdir_paths(WORK_DIR):
            fuzzer, target, repeat = utility.parse_path_by(ar_path)
            bug_info.setdefault(target, [])
            plot_data_path = utility.search_file(ar_path, PLOT_DATA)
            assert plot_data_path, f"{plot_data_path} not exists"
            crash_to_time = BugFoundByTimeAnalysis.get_csv_data(plot_data_path)
            reports_unique_line_path = os.path.join(
                WORK_DIR,
                TRIAGE_BY_CASR,
                fuzzer,
                target,
                repeat,
                "reports_unique_line",
            )
            if os.path.exists(reports_unique_line_path):
                for report in os.listdir(reports_unique_line_path):
                    if not report.endswith(".casrep"):
                        continue
                    with open(os.path.join(reports_unique_line_path, report), "r") as f:
                        json_data = json.load(f)
                    bug_type = json_data["CrashSeverity"]["ShortDescription"]
                    crash_line = json_data["CrashLine"].split("/")[-1]
                    id = int(report.split(",")[0].lstrip("id:"))
                    id += 1
                    for crash_index in sorted(crash_to_time.keys()):
                        if crash_index >= id:
                            bug_found_time = crash_to_time[crash_index]
                            break
                    assert bug_found_time, "IMPOSSIBLE!! no crash time found"
                    bug_info[target].append(
                        {
                            FUZZER: fuzzer,
                            REPEAT: repeat,
                            "bug_found_time": utility.get_readable_time(bug_found_time),
                            "bug_type": bug_type,
                            "crash_line": crash_line,
                            "bug_field": bug_type + "/" + crash_line,
                        }
                    )
            else:
                bug_info[target].append(
                    {
                        FUZZER: fuzzer,
                        REPEAT: repeat,
                    }
                )
        return bug_info

    @staticmethod
    @utility.time_count("SAVE BUG FOUND BY TIME DONE!")
    def save(WORK_DIR, OUTPUT_FILE=None):
        bug_info = BugFoundByTimeAnalysis.obtain(WORK_DIR)
        if OUTPUT_FILE is None:
            OUTPUT_FILE = os.path.join(
                WORK_DIR,
                f"{os.path.basename(WORK_DIR)}_bug_found_by_time.xlsx",
            )
        excel_manager = ExcelManager()
        for target in sorted(bug_info.keys()):
            table_data = bug_info[target]
            bug_fields = sorted(
                list(
                    set(
                        [
                            item["bug_field"]
                            for item in table_data
                            if "bug_field" in item
                        ]
                    )
                ),
                key=CasrTriageAnalysis.sort_by_severity_and_crashline,
            )
            display_fields = [
                FUZZER,
                REPEAT,
            ] + bug_fields
            excel_manager.create_sheet(target)
            # the header of table
            excel_manager.set_sheet_header(
                display_fields + [TOTAL_BUGS],
                [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                            color=CasrTriageAnalysis.get_severity_color(
                                display_field.split("/", 1)[0].strip()
                            ),
                        )
                    }
                    for display_field in display_fields
                ]
                + [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                        )
                    }
                ],
                direction=VERTICAL,
            )
            tmp_data = {}
            for item in table_data:
                tmp_data.setdefault(item[FUZZER], {}).setdefault(item[REPEAT], {})
                if "bug_field" not in item:
                    continue
                tmp_data[item[FUZZER]][item[REPEAT]][item["bug_field"]] = item[
                    "bug_found_time"
                ]
            table_data = [
                {FUZZER: fuzzer, REPEAT: repeat, **tmp_data[fuzzer][repeat]}
                for fuzzer in tmp_data
                for repeat in tmp_data[fuzzer]
            ]
            table_data = sorted(
                table_data,
                key=lambda x: (
                    len(x) - 2,  # order by the number of bugs
                    x[FUZZER],
                    x[REPEAT],
                ),
                reverse=True,
            )
            # the rows of table
            for item in table_data:
                excel_manager.set_sheet_data(
                    [
                        item[display_field] if display_field in item.keys() else ""
                        for display_field in display_fields
                    ]
                    + [len(item) - 2],
                    [
                        {
                            "Fill": PatternFill(
                                fgColor=CasrTriageAnalysis.fuzzer_colors[item[FUZZER]],
                                fill_type="solid",
                            )
                        }
                        if item[FUZZER] in CasrTriageAnalysis.fuzzer_colors.keys()
                        else {}
                        for _ in display_fields
                    ]
                    + [{}],
                    direction=VERTICAL,
                )
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(f"Bug found by time are saved in {OUTPUT_FILE}")
        return OUTPUT_FILE

    @staticmethod
    def merge(
        EXCEL_PATHS, OUTPUT_FILE, HEAP_RELATED_BUGS_FIELD=False, TRIAGE_RULE=None
    ):
        assert OUTPUT_FILE is not None, f"{OUTPUT_FILE} is None"
        targets = set()
        for excel_path in EXCEL_PATHS:
            assert os.path.exists(excel_path), f"{excel_path} not exists"
            wb = openpyxl.load_workbook(excel_path)
            targets.update(wb.sheetnames)
        excel_manager = ExcelManager()
        for target in sorted(list(targets)):
            excels = {}
            repeats = {}
            for excel_path in EXCEL_PATHS:
                wb = openpyxl.load_workbook(excel_path)
                if target not in wb.sheetnames:
                    continue
                sheet = wb[target]
                header_fields = [
                    cell
                    for row in sheet.iter_cols(min_col=1, max_col=1, values_only=True)
                    for cell in row
                ]
                excel = None
                # get sheet data
                for data_col in sheet.iter_cols(min_col=2, values_only=True):
                    for header_field, data_value in zip(header_fields, data_col):
                        if header_field == FUZZER:
                            excel = excels.setdefault(data_value, [])
                            repeats.setdefault(data_value, 0)
                            repeats[data_value] += 1
                        break
                    else:
                        assert (
                            excel is not None
                        ), f"fuzzer not found in the first column of {target} in {excel_path}"
                    excel.append(
                        {
                            header_field: data_value
                            for header_field, data_value in zip(header_fields, data_col)
                            if data_value is not None and "/" in header_field
                        }
                    )
            display_fields = set()
            sheet_data = []
            # flatten data
            for fuzzer, total in excels.items():
                tmp_dict = {}
                average = 0
                m_average = 0
                has_seen = set()
                for item in total:
                    for header_field, data_value in item.items():
                        if TRIAGE_RULE:
                            if TRIAGE_RULE[target][header_field] not in has_seen:
                                has_seen.add(TRIAGE_RULE[target][header_field])
                                average += 1
                                if CasrTriageAnalysis.is_heap_related_bug(
                                    header_field.split("/")[0].strip()
                                ):
                                    m_average += 1
                            header_field = TRIAGE_RULE[target][header_field]
                        else:
                            average += 1
                            if CasrTriageAnalysis.is_heap_related_bug(
                                header_field.split("/")[0].strip()
                            ):
                                m_average += 1
                        display_fields.add(header_field)
                        if header_field not in tmp_dict:
                            tmp_dict[header_field] = data_value
                            continue
                        bug_found_time = utility.human_readable_to_timedelta(
                            data_value.strip()
                        )
                        pre_bug_found_time = utility.human_readable_to_timedelta(
                            tmp_dict[header_field].strip()
                        )
                        if bug_found_time < pre_bug_found_time:
                            tmp_dict[header_field] = data_value
                    has_seen.clear()
                # print(target, fuzzer, m_average, average,repeats[fuzzer])
                average /= repeats[fuzzer]
                m_average /= repeats[fuzzer]
                tmp_dict[TOTAL_BUGS] = (
                    str(len(tmp_dict.keys()))
                    + "/"
                    + str(round(average, 2))
                    + "/"
                    + str(repeats[fuzzer])
                )
                tmp_dict[FUZZER] = fuzzer
                if HEAP_RELATED_BUGS_FIELD:
                    memory_related_bugs = sum(
                        [
                            1
                            for field in tmp_dict.keys()
                            if field != FUZZER
                            and field != TOTAL_BUGS
                            and CasrTriageAnalysis.is_heap_related_bug(
                                field.split("/")[0].strip()
                            )
                        ]
                    )
                    tmp_dict[HEAP_RELATED_BUGS] = (
                        str(memory_related_bugs)
                        + "/"
                        + str(round(m_average, 2))
                        + "/"
                        + str(repeats[fuzzer])
                    )
                    assert (
                        memory_related_bugs <= len(tmp_dict.keys()) - 2
                    ), "This should not happen"
                sheet_data.append(tmp_dict)
            display_fields = sorted(
                list(display_fields),
                key=CasrTriageAnalysis.sort_by_severity_and_crashline,
            )
            display_fields = [FUZZER] + display_fields
            if HEAP_RELATED_BUGS_FIELD:
                display_fields += [HEAP_RELATED_BUGS]
            display_fields += [TOTAL_BUGS]
            excel_manager.create_sheet(target)
            excel_manager.set_sheet_header(
                display_fields,
                [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                            color=CasrTriageAnalysis.get_severity_color(
                                display_field.split("/")[0].strip()
                            ),
                        )
                    }
                    for display_field in display_fields
                ],
                direction=VERTICAL,
            )
            for data in sorted(
                sheet_data,
                key=lambda x: (
                    int(x[TOTAL_BUGS].split("/")[0]),
                    float(x[TOTAL_BUGS].split("/")[1]),
                    int(x[TOTAL_BUGS].split("/")[2]),
                    x[FUZZER],
                ),
                reverse=True,
            ):
                row_data = []
                row_style = []
                for display_field in display_fields:
                    if display_field not in data:
                        row_data.append("")
                    else:
                        row_data.append(data[display_field])
                    if data[FUZZER] in BugFoundByTimeAnalysis.fuzzer_colors:
                        row_style.append(
                            {
                                "Fill": PatternFill(
                                    fgColor=BugFoundByTimeAnalysis.fuzzer_colors[
                                        data[FUZZER]
                                    ],
                                    fill_type="solid",
                                )
                            }
                        )
                    else:
                        row_style.append({})
                excel_manager.set_sheet_data(row_data, row_style, direction=VERTICAL)
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(
            f"The merged bug found by time result can be found in {OUTPUT_FILE}"
        )
