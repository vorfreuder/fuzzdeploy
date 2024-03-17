import os
import re
import time

from openpyxl.styles import Font, PatternFill
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TextColumn,
    TimeElapsedColumn,
)

from . import utility
from .Builder import Builder
from .constants import *
from .CpuAllocator import CpuAllocator
from .ExcelManager import ExcelManager
from .Maker import Maker


class CasrTriageAnalysis:
    fuzzer_colors = {
        "aflplusplus": "00A0B0",
        "htfuzz": "FBD26A",
    }
    display_fields = [
        FUZZER,
        REPEAT,
        "unique_line",
        "casr_dedup",
        "casr_dedup_cluster",
    ]
    Severity = {
        "EXPLOITABLE": (
            "SegFaultOnPc",
            "ReturnAv",
            "BranchAv",
            "CallAv",
            "DestAv",
            "BranchAvTainted",
            "CallAvTainted",
            "DestAvTainted",
            "heap-buffer-overflow(write)",
            "global-buffer-overflow(write)",
            "stack-use-after-scope(write)",
            "stack-use-after-return(write)",
            "stack-buffer-overflow(write)",
            "stack-buffer-underflow(write)",
            "heap-use-after-free(write)",
            "container-overflow(write)",
            "param-overlap",
        ),
        "PROBABLY_EXPLOITABLE": (
            "BadInstruction",
            "SegFaultOnPcNearNull",
            "BranchAvNearNull",
            "CallAvNearNull",
            "HeapError",
            "StackGuard",
            "DestAvNearNull",
            "heap-buffer-overflow",
            "global-buffer-overflow",
            "stack-use-after-scope",
            "use-after-poison",
            "stack-use-after-return",
            "stack-buffer-overflow",
            "stack-buffer-underflow",
            "heap-use-after-free",
            "container-overflow",
            "negative-size-param",
            "calloc-overflow",
            "readllocarray-overflow",
            "pvalloc-overflow",
            "overwrites-const-input",
        ),
        "NOT_EXPLOITABLE": (
            "SourceAv",
            "AbortSignal",
            "AccessViolation",
            "SourceAvNearNull",
            "SafeFunctionCheck",
            "FPE",
            "StackOverflow",
            "double-free",
            "bad-free",
            "alloc-dealloc-mismatch",
            "heap-buffer-overflow(read)",
            "global-buffer-overflow(read)",
            "stack-use-after-scope(read)",
            "stack-use-after-return(read)",
            "stack-buffer-overflow(read)",
            "stack-buffer-underflow(read)",
            "heap-use-after-free(read)",
            "container-overflow(read)",
            "initialization-order-fiasco",
            "new-delete-type-mismatch",
            "bad-malloc_usable_size",
            "odr-violation",
            "memory-leaks",
            "invalid-allocation-alignment",
            "invalid-aligned-alloc-alignment",
            "invalid-posix-memalign-alignment",
            "allocation-size-too-big",
            "out-of-memory",
            "fuzz target exited",
            "timeout",
        ),
    }

    @staticmethod
    def is_heap_related_bug(bug_type):
        if "stack" in bug_type.lower():
            return False
        return bug_type not in (
            "param-overlap",
            "BadInstruction",
            "overwrites-const-input",
            "AbortSignal",
            "SafeFunctionCheck",
            "FPE",
            "initialization-order-fiasco",
            "odr-violation",
            "fuzz target exited",
            "timeout",
        )

    @staticmethod
    def is_triaged_by_casr(fuzzer, target, repeat, dst_path, work_dir):
        if not os.path.exists(dst_path):
            return False
        ar_path = os.path.join(work_dir, "ar", fuzzer, target, repeat)
        crash_path = utility.search_item(ar_path, "FOLDER", "crashes")
        crash_set = set(os.listdir(crash_path))
        crash_set.discard("README.txt")
        if (len(crash_set) == 0) and (
            not os.path.exists(os.path.join(dst_path, "summary_by_unique_line"))
        ):
            return False
        for folder_name in ["failed", "reports"]:
            path = os.path.join(dst_path, folder_name)
            if os.path.exists(path):
                for file_name in os.listdir(path):
                    if folder_name == "reports":
                        if not file_name.endswith(".casrep"):
                            continue
                        file_name = file_name.rstrip(".casrep")
                    if file_name in crash_set:
                        crash_set.remove(file_name)
                    else:
                        os.remove(os.path.join(path, file_name))
        return len(crash_set) == 0

    @staticmethod
    @utility.time_count("CRASH TRIAGE BY CASR@https://github.com/ispras/casr DONE!")
    def obtain(WORK_DIR):
        total_crash_num = 0
        untriaged_paths = []
        for fuzzer, target, repeat, ar_path in utility.get_workdir_paths_by(WORK_DIR):
            fuzzer_stats_path = utility.search_item(ar_path, "FILE", FUZZER_STATS)
            if fuzzer_stats_path is None:
                utility.console.print(
                    f"[yellow]Warning: {FUZZER_STATS} not found in {ar_path}, maybe fine.[/yellow]"
                )
            crash_path = utility.search_item(ar_path, "FOLDER", "crashes")
            assert crash_path is not None, f"crashes not found in {ar_path}"
            crash_ls = os.listdir(crash_path)
            tmp_num = len(crash_ls)
            if "README.txt" in os.listdir(crash_path):
                tmp_num -= 1
            total_crash_num += tmp_num
            untriaged_paths.append((fuzzer, target, repeat))
        maker = Maker(
            WORK_DIR,
            TRIAGE_BY_CASR,
            "casr",
            IS_SKIP=CasrTriageAnalysis.is_triaged_by_casr,
            MODE="ALL",
        )
        with Progress(
            SpinnerColumn(spinner_name="arrow3"),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[bold]{task.completed} / {task.total}"),
            TimeElapsedColumn(),
            transient=True,
        ) as progress:
            triage_task = progress.add_task(
                "[bold green]Triaging", total=total_crash_num
            )
            last_triaged_crashes_num = 0
            while True:
                triaged_paths = []
                triaged_crashes_num = 0
                for fuzzer, target, repeat in untriaged_paths:
                    ar_path = os.path.join(WORK_DIR, "ar", fuzzer, target, repeat)
                    triaged_path = os.path.join(
                        WORK_DIR, TRIAGE_BY_CASR, fuzzer, target, repeat
                    )
                    if not os.path.exists(triaged_path):
                        continue
                    crash_path = utility.search_item(ar_path, "FOLDER", "crashes")
                    crash_set = set(os.listdir(crash_path))
                    crash_set.discard("README.txt")
                    triaged_crashes_num += len(crash_set)
                    for folder_name in ["failed", "reports"]:
                        path = os.path.join(triaged_path, folder_name)
                        if os.path.exists(path):
                            for file_name in os.listdir(path):
                                if folder_name == "reports":
                                    if not file_name.endswith(".casrep"):
                                        continue
                                    file_name = file_name.rstrip(".casrep")
                                if file_name in crash_set:
                                    crash_set.remove(file_name)
                    if len(crash_set) == 0:
                        triaged_paths.append(triaged_path)
                    triaged_crashes_num -= len(crash_set)
                untriaged_paths = list(set(untriaged_paths) - set(triaged_paths))
                progress.update(
                    triage_task,
                    advance=triaged_crashes_num - last_triaged_crashes_num,
                )
                last_triaged_crashes_num = triaged_crashes_num
                maker.thread.join(5)
                if not maker.thread.is_alive():
                    break
            progress.update(triage_task, completed=total_crash_num)
            # time.sleep(2)
        triage_results = {}
        for fuzzer, target, repeat, triaged_path in utility.get_workdir_paths_by(
            WORK_DIR, TRIAGE_BY_CASR
        ):
            triage = {
                FUZZER: fuzzer,
                REPEAT: repeat,
                "unique_line": 0,
                "casr_dedup": 0,
                "casr_dedup_cluster": 0,
            }
            if os.path.exists(os.path.join(triaged_path, "reports_dedup")):
                triage["casr_dedup"] = len(
                    os.listdir(os.path.join(triaged_path, "reports_dedup"))
                )
            if os.path.exists(os.path.join(triaged_path, "reports_dedup_cluster")):
                triage["casr_dedup_cluster"] = len(
                    os.listdir(os.path.join(triaged_path, "reports_dedup_cluster"))
                )
            summary_by_unique_line_path = os.path.join(
                triaged_path, "summary_by_unique_line"
            )
            if os.path.isfile(summary_by_unique_line_path):
                with open(summary_by_unique_line_path) as f:
                    summary_content = f.read()
                data_str = summary_content.split("->")[-1]
                pattern = r"(.+?): (\d+)"
                for match in re.findall(pattern, data_str):
                    key, value = match
                    key = key.strip()
                    value = int(value.strip())
                    triage[key] = value
                    triage["unique_line"] += value
            triage_results.setdefault(target, []).append(triage)
        for target in triage_results.keys():
            triage_results[target] = sorted(
                triage_results[target],
                key=lambda x: (
                    x["unique_line"],
                    x["casr_dedup_cluster"],
                    x["casr_dedup"],
                ),
                reverse=True,
            )
        utility.console.print(
            f"The results can be found in {os.path.join(WORK_DIR, TRIAGE_BY_CASR)}"
        )
        return triage_results

    @staticmethod
    def sort_by_severity(field):
        if field in CasrTriageAnalysis.Severity["EXPLOITABLE"]:
            return 0, field
        elif field in CasrTriageAnalysis.Severity["PROBABLY_EXPLOITABLE"]:
            return 1, field
        return 2, field

    @staticmethod
    def sort_by_severity_and_crashline(bug_field):
        field, crashline = bug_field.split("/", 1)
        return *CasrTriageAnalysis.sort_by_severity(field), crashline

    @staticmethod
    def get_severity_color(field):
        color = "000000"
        if field in CasrTriageAnalysis.Severity["EXPLOITABLE"]:
            color = "800000"
        elif field in CasrTriageAnalysis.Severity["PROBABLY_EXPLOITABLE"]:
            color = "1F497D"
        return color

    @staticmethod
    # @utility.time_count("SAVE TRIAGE BY CASR@https://github.com/ispras/casr DONE!")
    def save(WORK_DIR, OUTPUT_FILE=None, HEAP_RELATED_BUGS_FIELD=False):
        if OUTPUT_FILE is None:
            OUTPUT_FILE = os.path.join(
                WORK_DIR,
                f"{os.path.basename(WORK_DIR)}_triage_by_casr.xlsx",
            )
        triage_results = CasrTriageAnalysis.obtain(WORK_DIR)
        excel_manager = ExcelManager()
        for target in sorted(triage_results.keys()):
            table_data = triage_results[target]
            bug_fields = list(set([_ for item in table_data for _ in item.keys()]))
            bug_fields = [
                item
                for item in bug_fields
                if item not in CasrTriageAnalysis.display_fields
            ]
            bug_fields = sorted(bug_fields, key=CasrTriageAnalysis.sort_by_severity)
            if HEAP_RELATED_BUGS_FIELD:
                display_fields = (
                    CasrTriageAnalysis.display_fields + [HEAP_RELATED_BUGS] + bug_fields
                )
            else:
                display_fields = CasrTriageAnalysis.display_fields + bug_fields
            excel_manager.create_sheet(target)
            # the header of table
            excel_manager.set_sheet_header(
                display_fields,
                [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                            color=CasrTriageAnalysis.get_severity_color(display_field),
                        )
                    }
                    for display_field in display_fields
                ],
            )
            # the rows of table
            for item in table_data:
                if HEAP_RELATED_BUGS_FIELD:
                    item[HEAP_RELATED_BUGS] = 0
                    for bug_field in bug_fields:
                        if (
                            bug_field in item.keys()
                            and CasrTriageAnalysis.is_heap_related_bug(bug_field)
                        ):
                            item[HEAP_RELATED_BUGS] += item[bug_field]
                excel_manager.set_sheet_data(
                    [
                        item[display_field] if display_field in item.keys() else ""
                        for display_field in display_fields
                    ],
                    [
                        (
                            {
                                "Fill": PatternFill(
                                    fgColor=CasrTriageAnalysis.fuzzer_colors[
                                        item[FUZZER]
                                    ],
                                    fill_type="solid",
                                )
                            }
                            if item[FUZZER] in CasrTriageAnalysis.fuzzer_colors.keys()
                            else {}
                        )
                        for _ in display_fields
                    ],
                )
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(
            f"triage by casr@https://github.com/ispras/casr are saved in {OUTPUT_FILE}"
        )
        return OUTPUT_FILE
