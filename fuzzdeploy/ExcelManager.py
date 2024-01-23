import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class ExcelManager:
    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook["Sheet"])

    @staticmethod
    def realLength(string):
        """Calculate the byte length of a mixed single- and double-byte string"""
        dualByteNum = len("".join(re.compile("[^\x00-\xff]+").findall(string)))
        singleByteNum = len(string) - dualByteNum
        return (dualByteNum, singleByteNum, dualByteNum * 2 + singleByteNum)

    def create_sheet(self, sheet_name):
        self.sheet = self.workbook.create_sheet(title=sheet_name)

    def set_sheet_header(self, data, header_styles=None, direction="horizontal"):
        if header_styles is None:
            header_styles = [{} for _ in data]
        else:
            for index in range(len(header_styles)):
                if header_styles[index] is None:
                    header_styles[index] = {}
        assert len(data) == len(
            header_styles
        ), "data and header_styles should have the same length"
        for col_num, (header_text, header_style) in enumerate(
            zip(data, header_styles), 1
        ):
            if direction == "horizontal":
                cell = self.sheet.cell(row=1, column=col_num, value=header_text)
            else:
                cell = self.sheet.cell(row=col_num, column=1, value=header_text)
            cell.font = header_style.get(
                "Font",
                Font(
                    bold=True,
                    name="Calibri",
                    size=17,
                ),
            )
            cell.alignment = header_style.get(
                "Alignment", Alignment(horizontal="center", vertical="center")
            )
            cell.fill = header_style.get(
                "Fill",
                PatternFill(),
            )

    def set_sheet_data(self, data, data_styles=None, direction="horizontal"):
        if data_styles is None:
            data_styles = [{} for _ in data]
        else:
            for index in range(len(data_styles)):
                if data_styles[index] is None:
                    data_styles[index] = {}
        assert len(data) == len(
            data_styles
        ), "data and data_styles should have the same length"
        if direction == "horizontal":
            row_num = self.sheet.max_row + 1
        else:
            row_num = self.sheet.max_column + 1
        for col_num, (cell_value, cell_style) in enumerate(zip(data, data_styles), 1):
            if direction == "horizontal":
                cell = self.sheet.cell(row=row_num, column=col_num, value=cell_value)
            else:
                cell = self.sheet.cell(row=col_num, column=row_num, value=cell_value)
            cell.font = cell_style.get(
                "Font",
                Font(
                    name="Calibri",
                    size=17,
                ),
            )
            cell.alignment = cell_style.get(
                "Alignment", Alignment(horizontal="center", vertical="center")
            )
            cell.fill = cell_style.get(
                "Fill",
                PatternFill(),
            )

    def save_workbook(self, filename):
        # auto adjust width
        for sheet in self.workbook.worksheets:
            for col in sheet.columns:
                max_length = 0
                for cell in col:
                    try:
                        cell_len = ExcelManager.realLength(str(cell.value))[2]
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
                sheet.column_dimensions[col[0].column_letter].width = max_length * 1.7
        self.workbook.save(filename)
