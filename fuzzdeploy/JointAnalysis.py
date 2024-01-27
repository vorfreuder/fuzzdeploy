import copy
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

from . import utility
from .CasrTriageAnalysis import CasrTriageAnalysis
from .ExcelManager import ExcelManager
from .StateAnalysis import StateAnalysis
from .utility import MEMORY_RELATED_BUGS


class JointAnalysis:
    fuzzer_colors = {
        "aflplusplus": "00A0B0",
        "htfuzz": "FBD26A",
    }

    @staticmethod
    # @utility.time_count("MERGE FILES DONE!")
    def save_joint_results(STATS_FILE, TRIAGE_FILE, OUTPUT_FILE):
        assert os.path.exists(STATS_FILE), f"{STATS_FILE} not found"
        assert os.path.exists(TRIAGE_FILE), f"{TRIAGE_FILE} not found"
        assert OUTPUT_FILE is not None, f"{OUTPUT_FILE} is None"
        stats_wb = openpyxl.load_workbook(STATS_FILE)
        triage_wb = openpyxl.load_workbook(TRIAGE_FILE)
        excel_manager = ExcelManager()
        for triage_sheet in triage_wb.worksheets:
            sheet_name = triage_sheet.title
            assert (
                sheet_name in stats_wb.sheetnames
            ), f"{sheet_name} not found in {TRIAGE_FILE}. Is it generated from the same workdir?"
            stats_sheet = stats_wb[sheet_name]
            excel_manager.create_sheet(sheet_name)

            display_fields = list(triage_sheet[1])
            fuzzer_fields = list(stats_sheet[1])
            display_fields += [
                field
                for field in fuzzer_fields
                if field.value not in [_.value for _ in display_fields]
            ]
            excel_manager.set_sheet_header(
                [display_field.value for display_field in display_fields],
                [
                    {
                        "Font": copy.copy(display_field.font),
                        "Alignment": copy.copy(display_field.alignment),
                        "Fill": copy.copy(display_field.fill),
                    }
                    for display_field in display_fields
                ],
            )
            for triage_row in triage_sheet.iter_rows(min_row=2):
                fuzzer = triage_row[0].value
                repeat = triage_row[1].value
                for stats_row in stats_sheet.iter_rows(min_row=2):
                    if stats_row[0].value == fuzzer and stats_row[1].value == repeat:
                        row_data = []
                        row_style = []
                        for col_idx, cell in enumerate(
                            [cell for cell in triage_row] + list(stats_row[2:]), start=1
                        ):
                            row_data.append(cell.value)
                            row_style.append(
                                {
                                    "Font": copy.copy(cell.font),
                                    "Alignment": copy.copy(cell.alignment),
                                    "Fill": copy.copy(cell.fill),
                                }
                            )
                        break
                excel_manager.set_sheet_data(row_data, row_style)
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(f"The merged result can be found in {OUTPUT_FILE}")
        return OUTPUT_FILE

    @staticmethod
    def merge_joint_excels(EXCEL_PATHS, OUTPUT_FILE, MEMORY_RELATED_BUGS_FIELD=False):
        assert OUTPUT_FILE is not None, f"{OUTPUT_FILE} is None"
        excels = {}
        for excel_path in EXCEL_PATHS:
            assert os.path.exists(excel_path), f"{excel_path} not exists"
            wb = openpyxl.load_workbook(excel_path)
            for target in wb.sheetnames:
                sheet = wb[target]
                for data_row in sheet.iter_rows(min_row=2, values_only=True):
                    excel = {}
                    for header_field, data_value in zip(sheet[1], data_row):
                        excel[header_field.value] = data_value
                    if MEMORY_RELATED_BUGS_FIELD:
                        excel[MEMORY_RELATED_BUGS] = sum(
                            [
                                int(excel[field])
                                for field in excel.keys()
                                if field not in StateAnalysis.display_fields
                                and field not in CasrTriageAnalysis.display_fields
                                and field != MEMORY_RELATED_BUGS
                                and excel[field] is not None
                                and CasrTriageAnalysis.is_memory_related_bug(field)
                            ]
                        )
                    excels.setdefault(target, []).append(excel)
        excel_manager = ExcelManager()
        for target in sorted(excels.keys()):
            excel_manager.create_sheet(target)
            header = sorted(
                [
                    field
                    for field in list(
                        set(k for data_item in excels[target] for k in data_item.keys())
                    )
                    if field not in StateAnalysis.display_fields
                    and field not in CasrTriageAnalysis.display_fields
                    and field != MEMORY_RELATED_BUGS
                ],
                key=CasrTriageAnalysis.sort_by_severity,
            )
            if MEMORY_RELATED_BUGS_FIELD:
                display_fields = (
                    CasrTriageAnalysis.display_fields + [MEMORY_RELATED_BUGS] + header
                )
            else:
                display_fields = CasrTriageAnalysis.display_fields + header
            for field in StateAnalysis.display_fields:
                if field not in display_fields:
                    display_fields.append(field)
            excel_manager.set_sheet_header(
                display_fields,
                [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                            color=CasrTriageAnalysis.get_severity_color(display_field),
                        )
                    }
                    for display_field in display_fields
                ],
            )
            for data_item in sorted(
                excels[target],
                key=lambda x: (
                    x["unique_line"],
                    x["casr_dedup_cluster"],
                    x["casr_dedup"],
                    x["fuzzer"],
                ),
                reverse=True,
            ):
                row_data = []
                row_style = []
                for display_field in display_fields:
                    if display_field not in data_item:
                        row_data.append("")
                        row_style.append({})
                        continue
                    row_data.append(data_item[display_field])
                    if data_item["fuzzer"] in JointAnalysis.fuzzer_colors:
                        row_style.append(
                            {
                                "Fill": PatternFill(
                                    fgColor=JointAnalysis.fuzzer_colors[
                                        data_item["fuzzer"]
                                    ],
                                    fill_type="solid",
                                )
                            }
                        )
                    else:
                        row_style.append({})
                excel_manager.set_sheet_data(row_data, row_style)
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(f"The merged joint result can be found in {OUTPUT_FILE}")

    @staticmethod
    def merge_save_bug_found_by_speed_excels(
        EXCEL_PATHS, OUTPUT_FILE, MEMORY_RELATED_BUGS_FIELD=False, TRIAGE_RULE=None
    ):
        assert OUTPUT_FILE is not None, f"{OUTPUT_FILE} is None"
        targets = set()
        for excel_path in EXCEL_PATHS:
            assert os.path.exists(excel_path), f"{excel_path} not exists"
            wb = openpyxl.load_workbook(excel_path)
            targets.update(wb.sheetnames)
        excel_manager = ExcelManager()
        for target in sorted(list(targets)):
            excels = {}
            repeats = {}
            for excel_path in EXCEL_PATHS:
                wb = openpyxl.load_workbook(excel_path)
                if target not in wb.sheetnames:
                    continue
                sheet = wb[target]
                header_fields = [
                    cell
                    for row in sheet.iter_cols(min_col=1, max_col=1, values_only=True)
                    for cell in row
                ]
                excel = None
                # get sheet data
                for data_col in sheet.iter_cols(min_col=2, values_only=True):
                    for header_field, data_value in zip(header_fields, data_col):
                        if header_field == "fuzzer":
                            excel = excels.setdefault(data_value, [])
                            repeats.setdefault(data_value, 0)
                            repeats[data_value] += 1
                        break
                    else:
                        assert (
                            excel is not None
                        ), f"fuzzer not found in the first column of {target} in {excel_path}"
                    excel.append(
                        {
                            header_field: data_value
                            for header_field, data_value in zip(header_fields, data_col)
                            if data_value is not None
                            and "/" in header_field
                            and "/" in data_value
                        }
                    )
            display_fields = set()
            sheet_data = []
            # flatten data
            for fuzzer, total in excels.items():
                tmp_dict = {}
                average = 0
                m_average = 0
                has_seen = set()
                for item in total:
                    for header_field, data_value in item.items():
                        if TRIAGE_RULE:
                            if TRIAGE_RULE[target][header_field] not in has_seen:
                                has_seen.add(TRIAGE_RULE[target][header_field])
                                average += 1
                                if CasrTriageAnalysis.is_memory_related_bug(
                                    header_field.split("/")[0].strip()
                                ):
                                    m_average += 1
                            header_field = TRIAGE_RULE[target][header_field]
                        else:
                            average += 1
                            if CasrTriageAnalysis.is_memory_related_bug(
                                header_field.split("/")[0].strip()
                            ):
                                m_average += 1
                        display_fields.add(header_field)
                        if header_field not in tmp_dict:
                            tmp_dict[header_field] = data_value
                            continue
                        bug_found_time, _ = data_value.split("/")
                        bug_found_time = bug_found_time.strip()
                        pre_bug_found_time, _ = tmp_dict[header_field].split("/")
                        pre_bug_found_time = pre_bug_found_time.strip()
                        if (
                            bug_found_time == "unknown"
                            or pre_bug_found_time == "unknown"
                        ):
                            continue
                        bug_found_time = utility.human_readable_to_timedelta(
                            bug_found_time
                        )
                        pre_bug_found_time = utility.human_readable_to_timedelta(
                            pre_bug_found_time
                        )
                        if bug_found_time < pre_bug_found_time:
                            tmp_dict[header_field] = data_value
                    has_seen.clear()
                # print(target, fuzzer, m_average, average,repeats[fuzzer])
                average /= repeats[fuzzer]
                m_average /= repeats[fuzzer]
                tmp_dict["total_bugs"] = (
                    str(len(tmp_dict.keys()))
                    + "/"
                    + str(round(average, 2))
                    + "/"
                    + str(repeats[fuzzer])
                )
                tmp_dict["fuzzer"] = fuzzer
                if MEMORY_RELATED_BUGS_FIELD:
                    memory_related_bugs = sum(
                        [
                            1
                            for field in tmp_dict.keys()
                            if field != "fuzzer"
                            and field != "total_bugs"
                            and CasrTriageAnalysis.is_memory_related_bug(
                                field.split("/")[0].strip()
                            )
                        ]
                    )
                    tmp_dict[MEMORY_RELATED_BUGS] = (
                        str(memory_related_bugs)
                        + "/"
                        + str(round(m_average, 2))
                        + "/"
                        + str(repeats[fuzzer])
                    )
                    assert (
                        memory_related_bugs <= len(tmp_dict.keys()) - 2
                    ), "This should not happen"
                sheet_data.append(tmp_dict)
            display_fields = sorted(
                list(display_fields),
                key=CasrTriageAnalysis.sort_by_severity_and_crashline,
            )
            display_fields = ["fuzzer"] + display_fields
            if MEMORY_RELATED_BUGS_FIELD:
                display_fields += [MEMORY_RELATED_BUGS]
            display_fields += ["total_bugs"]
            excel_manager.create_sheet(target)
            excel_manager.set_sheet_header(
                display_fields,
                [
                    {
                        "Font": Font(
                            bold=True,
                            name="Calibri",
                            size=17,
                            color=CasrTriageAnalysis.get_severity_color(
                                display_field.split("/")[0].strip()
                            ),
                        )
                    }
                    for display_field in display_fields
                ],
                direction="vertical",
            )
            for data in sorted(
                sheet_data,
                key=lambda x: (
                    x["total_bugs"],
                    x["fuzzer"],
                ),
                reverse=True,
            ):
                row_data = []
                row_style = []
                for display_field in display_fields:
                    if display_field not in data:
                        row_data.append("")
                    else:
                        row_data.append(data[display_field])
                    if data["fuzzer"] in JointAnalysis.fuzzer_colors:
                        row_style.append(
                            {
                                "Fill": PatternFill(
                                    fgColor=JointAnalysis.fuzzer_colors[data["fuzzer"]],
                                    fill_type="solid",
                                )
                            }
                        )
                    else:
                        row_style.append({})
                excel_manager.set_sheet_data(row_data, row_style, direction="vertical")
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(
            f"The merged bug found by speed result can be found in {OUTPUT_FILE}"
        )
