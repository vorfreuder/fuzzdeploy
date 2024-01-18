import os

from openpyxl.styles import PatternFill

from . import utility
from .ExcelManager import ExcelManager


class StateAnalysis:
    fuzzer_colors = {
        "aflplusplus": "00A0B0",
        "htfuzz": "FBD26A",
    }
    display_fields = [
        "fuzzer",
        "repeat",
        "crashes",  # AFL++ not support
        "pending_favs",
        "execs_done",
        "bitmap_cvg",
        "execute_time",
        "execs_per_sec",
        "unique_hangs",  # AFL++ not support
        "pending_total",
        "execs_since_crash",
    ]

    @staticmethod
    def get_state_results(WORK_DIR):
        assert os.path.exists(WORK_DIR), f"{WORK_DIR} not exists"
        no_fuzzer_stats_ls = []
        state_results = {}
        for test_path in utility.test_paths(WORK_DIR):
            fuzzer, target, repeat = utility.parse_path_by(test_path)
            fuzzer_stats_path = utility.search_file(test_path, "fuzzer_stats")
            if fuzzer_stats_path is None:
                no_fuzzer_stats_ls.append(f"{fuzzer}/{target}/{repeat}")
                continue
            state = {"fuzzer": fuzzer, "repeat": repeat}
            with open(fuzzer_stats_path, "r") as f:
                fuzzer_stats = f.read()
            for item in fuzzer_stats.strip().split("\n"):
                l, r = item.split(":")
                state[l.strip()] = r.strip()
            div = (int(state["last_update"]) - int(state["start_time"])) / 3600
            state["execute_time"] = f"{div:.5f}"
            state["crashes"] = (
                int(state["unique_crashes"])
                if state.get("unique_crashes")
                else int(state["saved_crashes"])
            )
            state["unique_hangs"] = (
                int(state["unique_hangs"])
                if state.get("unique_hangs")
                else int(state["saved_hangs"])
            )
            state_results.setdefault(target, []).append(state)
        for target in state_results.keys():
            state_results[target] = sorted(
                state_results[target],
                key=lambda x: int(x["crashes"]) / float(x["execute_time"]),
                reverse=True,
            )
        assert (
            len(state_results) > 0
        ), "No fuzzer_stats found! Is fuzzer working correctly?"
        if len(no_fuzzer_stats_ls) > 0:
            utility.console.print(
                f"[yellow]Warning: {len(no_fuzzer_stats_ls)} fuzzer_stats not found:[/yellow]",
                end=" ",
            )
            for item in no_fuzzer_stats_ls:
                print(f"{item}", end=" ")
            utility.console.print(f"[yellow]and ignored in output file![/yellow]")
        return state_results

    @staticmethod
    # @utility.time_count("SAVE FUZZER_STATS DONE!")
    def save_state_results(WORK_DIR, OUTPUT_FILE=None):
        if OUTPUT_FILE is None:
            OUTPUT_FILE = os.path.join(
                os.path.dirname(WORK_DIR),
                f"{os.path.basename(WORK_DIR)}_fuzzer_stats.xlsx",
            )
        state_results = StateAnalysis.get_state_results(WORK_DIR)
        excel_manager = ExcelManager()
        display_fields = StateAnalysis.display_fields
        for target in sorted(state_results.keys()):
            table_data = state_results[target]
            excel_manager.create_sheet(target)
            # the header of table
            excel_manager.set_sheet_header(display_fields)
            # the rows of table
            for item in table_data:
                excel_manager.set_sheet_data(
                    [
                        item[display_field] if display_field in item.keys() else ""
                        for display_field in display_fields
                    ],
                    [
                        {
                            "Fill": PatternFill(
                                fgColor=StateAnalysis.fuzzer_colors[item["fuzzer"]],
                                fill_type="solid",
                            )
                        }
                        if item["fuzzer"] in StateAnalysis.fuzzer_colors.keys()
                        else {}
                        for _ in display_fields
                    ],
                )
        excel_manager.save_workbook(OUTPUT_FILE)
        utility.console.print(f"The result can be found in {OUTPUT_FILE}")
        return OUTPUT_FILE
